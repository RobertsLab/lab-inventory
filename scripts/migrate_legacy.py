#!/usr/bin/env python3
"""One-time migration: legacy/Roberts Lab Inventory.xlsx -> data/*.csv

Emits three files:
  data/locations.csv     controlled vocabulary of physical places
  data/items.csv         one row per item, canonical inventory
  data/review_queue.csv  rows a human needs to look at before they become items

Design rule: never guess silently. Where the legacy data is ambiguous
(comma-delimited contents, suspected duplicates, unparseable dates) the row
goes to review_queue.csv with a stated reason instead of into items.csv.

Usage:
    python3 scripts/migrate_legacy.py [--xlsx PATH] [--out DIR]
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from collections import Counter
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install -r requirements.txt")

# The spreadsheet's own claim about when it was last updated (sheet names say
# 82921 / 9521 -> Aug 29 2021 / Sep 5 2021). Every migrated item inherits this
# as last_verified so the staleness is queryable rather than a footnote.
LEGACY_VERIFIED = "2021-08-29"
LEGACY_SOURCE = "legacy-xlsx"

ITEM_FIELDS = [
    "item_id", "name", "category", "location_id", "quantity", "unit",
    "vendor", "catalog_no", "lot", "received", "expires", "status",
    "last_verified", "verified_by", "owner", "source", "photo_id", "notes",
]
LOCATION_FIELDS = [
    "location_id", "room", "kind", "number", "parent_id", "label", "notes",
]
REVIEW_FIELDS = [
    "reason", "sheet", "row", "location_id", "raw_text", "suggestion", "notes",
]

# Rooms get short codes so location_ids stay ID-safe. "-80˚C room" would make
# a hostile identifier.
ROOM_CODES = {
    "209": "209",
    "213": "213",
    "230": "230",
    "-80˚c room": "M80",
}

KIND_MAP = {
    "cabinet": "cabinet",
    "cabinets": "cabinet",
    "drawer": "drawer",
    "shelf": "shelf",
    "shelves": "shelf",
    "refrigerator": "refrigerator",
    "-20 freezer": "freezer",
    "storage bin": "bin",
    "black shelving unit": "shelf",
}

# Short, fixed codes keep location_ids compact. These end up in QR payloads and
# URL fragments, so they must stay short and stable -- never derive them by
# truncating the kind name.
KIND_CODES = {
    "cabinet": "CAB", "drawer": "DRW", "shelf": "SH", "bin": "BIN",
    "refrigerator": "FRIDGE", "freezer": "FRZ", "bench": "BNCH",
    "floor": "FLR", "other": "OTH",
}

# Ordinal shelf labels used inconsistently across sheets ("Top", "2nd", ...).
ORDINALS = {
    "top": 1, "1st": 1, "2nd": 2, "3rd": 3, "4th": 4, "5th": 5,
    "6th": 6, "bottom": 99,
}

# Order matters: first match wins. Tuned against the legacy contents so that
# the site's category filter is actually useful; ~1 in 6 items still lands in
# `other`, which is honest for a drawer containing a Nintendo.
CATEGORY_RULES = [
    ("kit", r"\bkits?\b|mini ?prep|midi ?prep|maxi ?prep|micro ?prep|\bduet\b"),
    ("enzyme", r"polymerase|ligase|transcriptase|nuclease|dnase|rnase|proteinase|"
               r"restriction enzyme|\btaq\b|exonuclease|phosphatase|lysozyme|"
               r"\benzymes?\b"),
    ("antibody", r"antibod|antisera|\bigg\b|\bighm\b|\bserum\b|immunoglobulin"),
    ("media", r"\bagar\b|\bbroth\b|\bmedia\b|\bmedium\b|\bfbs\b|\bsera\b|"
              r"\bfood\b|algae paste"),
    ("sample", r"\bsamples?\b|\bgdna\b|extractions?|\baliquots?\b|libraries|"
               r"snap cap|histology slides|\bcdna\b|\brna\b(?!ase)\s+from|"
               r"\bblood\b|tissue in|in 100% ethanol|\bbcs plates?\b"),
    ("glassware", r"\bbottles?\b|\bflasks?\b|\bbeakers?\b|graduated cylinders?|"
                  r"\bjars?\b|\bcuvettes?\b|cover ?slips?|\bpetri\b|"
                  r"weigh (boats?|paper)|\bpipette? bulbs?\b|\bglassware\b"),
    ("equipment", r"centrifuge|thermocycler|shaker|manifold|incubator|"
                  r"spectrophotometer|\bbalance\b|\bpump\b|magnet|dynamag|"
                  r"nucleomag|\bfridge\b|freezer|\blamps?\b|light ?bulbs?|"
                  r"satellite tags?|mac mini|external drives?|\bhdd\b|"
                  r"web ?cam|keyboards?|computer (mouse|mice)|monitor|"
                  r"microscope|cellometers?|homogenizer|pestle|electrode|"
                  r"pipettors?|repeater pipets?|\bsonicator\b|hot ?plate|"
                  r"vortex|nanodrop|\bqubit\b(?! (rna|dna))"),
    ("tool", r"forceps|tweezers|scissors|scalpel|wrench|screwdriver|pliers|"
             r"\btools?\b|zip ties|\bclamps?\b|spatulas?|\bscoops?\b|"
             r"\bbrush(es)?\b|paintbrush|\bsiphon\b|thermometer|\btimer\b|"
             r"stir bars?|\bfunnels?\b|\bprobe\b|\bpaddles?\b|\bkeys?\b|"
             r"switchblade|\bmesh\b|razor"),
    ("consumable", r"\btips?\b|\btubes?\b|\bplates?\b|gloves|wipes|\bfoil\b|"
                   r"parafilm|syringes?|needles?|\bcolumns?\b|\bslides?\b|"
                   r"pipets?\b|pipettes?\b|dishes|\bboxes\b|\bbags?\b|"
                   r"\bracks?\b|\bfilters?\b|plastic wrap|\bcaps?\b|"
                   r"\bcontainers?\b|cassettes?|\bchips?\b|\bstrips?\b|"
                   r"benchtop protectors|\bswabs?\b|\bvials?\b|"
                   r"incubation chambers|\bsheaths?\b|\bcords?\b|batteries"),
    ("reagent", r"buffers?|reagants?|reagents?|ethanol|methanol|isopropanol|"
                r"\btris\b|\bedta\b|\bsds\b|\bacid\b|\bnaoh\b|\bhcl\b|"
                r"\bstain\b|\bdyes?\b|\bladder\b|primers?|\bdntp|\bbsa\b|"
                r"glycerol|agarose|acrylamide|\bresin\b|\bbeads?\b|trizol|"
                r"tri reagant|substrate|\bsolutions?\b|\bassay\b|rnazol|"
                r"\bcontrol dna\b|\bstandards?\b|\bwater\b|\bsalts?\b|"
                r"formalin|formaldehyde|\bpbs\b|\bdmso\b"),
    ("office", r"\bpens?\b|\bpencils?\b|sharpies?|\bstaples?\b|post-?its?|"
               r"\bpaper\b|notebooks?|composition book|file folders?|"
               r"white-?out|\bglue\b|paper clips?|\btacks?\b|stickers?|"
               r"\btape\b|report covers?|clipboard|calculator|handbooks?|"
               r"manuals?|business card|tough-?spots?|\bl?ab?els?\b|"
               r"command strips|\bbinders?\b|\bnotepad\b|\bmarkers?\b"),
]

# " - Goetz" / " -Goetz" trailing owner attribution used in the storage bins.
OWNER_RE = re.compile(r"\s+-\s*([A-Z][A-Za-z'-]+)\s*$")
# "(11/27/2019)" embedded date
PAREN_DATE_RE = re.compile(r"\((\d{1,2})/(\d{1,2})/(\d{2,4})\)")
# "(x2)" / "(X6)" multiplicity
XQTY_RE = re.compile(r"\(\s*[xX]\s*(\d+)\s*\)")
# "In cardboard box labeled "HOLD": a, b, c" container prefix
PREFIX_RE = re.compile(r'^([^;:]{4,70}):\s+(.*\S)$')


def slug(text: str) -> str:
    """Turn an arbitrary label into an ID-safe fragment."""
    s = re.sub(r"[^A-Za-z0-9]+", "-", str(text).strip()).strip("-").upper()
    return re.sub(r"-{2,}", "-", s)


def clean(value) -> str:
    """Normalize a cell to a trimmed string. 209.0 -> '209', 7.0 -> '7'."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    return str(value).strip()


def room_code(raw: str) -> str:
    key = clean(raw).lower()
    return ROOM_CODES.get(key, slug(key) or "UNK")


def norm_kind(raw: str) -> str:
    key = clean(raw).lower().rstrip(" .")
    return KIND_MAP.get(key, "other")


DATE_FORMATS = ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%m/%Y", "%Y-%m", "%Y")
PARTIAL_LEN = {"%m/%Y": 7, "%Y-%m": 7, "%Y": 4}


def parse_date(value, sheet: str, row: int, warnings: list) -> tuple[str, str]:
    """Parse a legacy date cell.

    Returns (iso_date, residue). Partial dates keep their precision
    ('08/2010' -> '2010-08') rather than being invented into a day. Text that
    isn't a date at all ('Very likely all expired', 'n.d.') is returned as
    residue for the caller to preserve in notes -- the legacy sheet used the
    date column for commentary and that commentary is worth keeping.
    """
    if value in (None, ""):
        return "", ""
    if isinstance(value, dt.datetime):
        d, precision = value.date(), 10
    else:
        text = clean(value)
        # Leading date with trailing commentary: '08/2010 from FHL'
        residue = ""
        head = text
        match = re.match(r"^([\d/.-]+)\s+(.*\S)$", text)
        if match:
            head, residue = match.group(1), match.group(2)
        for fmt in DATE_FORMATS:
            try:
                d = dt.datetime.strptime(head, fmt).date()
                precision = PARTIAL_LEN.get(fmt, 10)
                break
            except ValueError:
                continue
        else:
            warnings.append((sheet, row, f"date column held non-date text: {text!r}"
                             " -> moved to notes"))
            return "", text
        if residue:
            warnings.append((sheet, row,
                             f"kept date {head!r}, moved {residue!r} to notes"))
        if not (dt.date(1990, 1, 1) <= d <= dt.date.today()):
            warnings.append((sheet, row, f"implausible date: {d.isoformat()}"))
        return d.isoformat()[:precision], residue

    if not (dt.date(1990, 1, 1) <= d <= dt.date.today()):
        warnings.append((sheet, row, f"implausible date: {d.isoformat()}"))
    return d.isoformat()[:precision], ""


def categorize(name: str) -> str:
    low = name.lower()
    for category, pattern in CATEGORY_RULES:
        if re.search(pattern, low):
            return category
    return "other"


class Migration:
    def __init__(self, xlsx: Path, outdir: Path):
        self.wb = openpyxl.load_workbook(xlsx, data_only=True)
        self.outdir = outdir
        self.locations: dict[str, dict] = {}
        self.items: list[dict] = []
        self.review: list[dict] = []
        self.warnings: list[tuple] = []
        self.skipped = 0
        self._next_id = 1

    # -- registries ---------------------------------------------------------

    def location(self, location_id: str, room: str, kind: str, number: str = "",
                 parent_id: str = "", label: str = "", notes: str = "") -> str:
        existing = self.locations.get(location_id)
        if existing:
            if notes and not existing["notes"]:
                existing["notes"] = notes
            return location_id
        self.locations[location_id] = {
            "location_id": location_id, "room": room, "kind": kind,
            "number": number, "parent_id": parent_id, "label": label,
            "notes": notes,
        }
        return location_id

    def add_item(self, name: str, location_id: str, **kw) -> None:
        name = name.strip(" ;,.")
        if not name:
            return
        row = {f: "" for f in ITEM_FIELDS}
        row.update({
            "item_id": f"itm-{self._next_id:05d}",
            "name": name,
            "category": categorize(name),
            "location_id": location_id,
            "status": "unverified",
            "last_verified": LEGACY_VERIFIED,
            "source": LEGACY_SOURCE,
        })
        row.update({k: v for k, v in kw.items() if v not in (None, "")})
        self._next_id += 1
        self.items.append(row)

    def flag(self, reason: str, sheet: str, row: int, location_id: str,
             raw_text: str, suggestion: str = "", notes: str = "") -> None:
        self.review.append({
            "reason": reason, "sheet": sheet, "row": row,
            "location_id": location_id, "raw_text": raw_text,
            "suggestion": suggestion, "notes": notes,
        })

    # -- content parsing ----------------------------------------------------

    def explode(self, contents: str, location_id: str, sheet: str, rownum: int,
                base_notes: str = "", received: str = "") -> None:
        """Split a legacy 'Contents' cell into items.

        Semicolons are a reliable delimiter in this data and are split
        automatically. Commas are not (`Microcentrifuge tubes, conical screw
        top tubes` is two items; `DNeasy 96 blood and tissue kit (4) & (50)` is
        one), so comma-only cells become a single item plus a review_queue
        entry carrying the suggested split.
        """
        text = contents.strip()
        if not text:
            return

        if text.lower().strip(" .") in ("empty", "none", "nothing"):
            loc = self.locations.get(location_id)
            if loc is not None and not loc["notes"]:
                loc["notes"] = f"recorded empty as of {LEGACY_VERIFIED}"
            return

        notes = base_notes

        # "In cardboard box labeled "HOLD": a, b, c" -> container becomes a note
        prefix_match = PREFIX_RE.match(text)
        if prefix_match and prefix_match.group(2).count(",") >= 2:
            container, text = prefix_match.group(1).strip(), prefix_match.group(2)
            notes = f"{notes}; {container}".strip("; ") if notes else container

        parts = [p.strip() for p in text.split(";") if p.strip()]
        multi = len(parts) > 1

        for part in parts:
            if not multi and part.count(",") >= 2:
                suggestion = " | ".join(
                    s.strip() for s in part.split(",") if s.strip())
                self.flag("comma-delimited-contents", sheet, rownum,
                          location_id, part, suggestion,
                          "kept as one item; split by hand if these are distinct")

            item_notes = notes
            item_received = received
            name = part

            date_match = PAREN_DATE_RE.search(name)
            if date_match:
                month, day, year = (int(g) for g in date_match.groups())
                year += 2000 if year < 100 else 0
                try:
                    item_received = dt.date(year, month, day).isoformat()
                except ValueError:
                    self.warnings.append(
                        (sheet, rownum, f"bad embedded date in {part!r}"))
                name = PAREN_DATE_RE.sub("", name).strip()

            quantity = ""
            qty_match = XQTY_RE.search(name)
            if qty_match:
                quantity = qty_match.group(1)
                name = XQTY_RE.sub("", name).strip()

            owner = ""
            owner_match = OWNER_RE.search(name)
            if owner_match:
                owner = owner_match.group(1)
                name = OWNER_RE.sub("", name).strip()
            elif name.strip() in ("Goetz",):
                owner, name = name.strip(), "unspecified contents"

            self.add_item(name, location_id, quantity=quantity, owner=owner,
                          received=item_received, notes=item_notes)

    # -- sheet handlers -----------------------------------------------------

    def rows(self, sheet_name: str):
        ws = self.wb[sheet_name]
        for rownum, values in enumerate(ws.iter_rows(min_row=2, values_only=True),
                                        start=2):
            if any(v not in (None, "") for v in values):
                yield rownum, values

    def drawers_cabinets(self, sheet: str) -> None:
        """Room / Drawer-Cabinet / Number / (unnamed contents column)."""
        # These free-text 'kind' values are coarse restatements of dedicated
        # sheets; their contents duplicate the per-item rows there.
        crossref_kinds = {
            "shelves to the left of entry-way (reference other tab)": "213-SHELF",
        }
        unnamed_seq: Counter = Counter()

        for rownum, values in self.rows(sheet):
            room_raw, kind_raw, number_raw, contents_raw = (
                clean(values[0]), clean(values[1]), clean(values[2]),
                clean(values[3]) if len(values) > 3 else "")
            room = room_code(room_raw)
            kind_key = kind_raw.lower().rstrip(" .")
            kind = norm_kind(kind_raw)

            # Pure cross-references to other sheets — no data of their own.
            if contents_raw.lower().startswith("see other tab"):
                self.skipped += 1
                continue

            # Coarse duplicates of a dedicated sheet: send contents to review
            # rather than creating ~40 duplicate items.
            if kind_key in crossref_kinds:
                canonical = crossref_kinds[kind_key]
                self.flag("coarse-duplicate-of-dedicated-sheet", sheet, rownum,
                          canonical, contents_raw,
                          " | ".join(p.strip() for p in contents_raw.split(";")),
                          f"shelf {number_raw!r}; reconcile against {canonical}-* "
                          "rows before adding")
                continue

            # Build the location id. The 'Number' column is sometimes a real
            # number, sometimes an ordinal ('Top', '2nd'), and sometimes prose
            # ('Long clear bin on top of VWR fridge (1)'). Prose gets a
            # sequential id and keeps its text in `label` -- slugifying it
            # produced 60-character ids that are hostile as QR payloads.
            code = KIND_CODES.get(kind, "OTH")
            ordinal = ORDINALS.get(number_raw.lower())
            if number_raw.isdigit():
                suffix, label = f"{int(number_raw):02d}", f"{kind_raw.strip()} {number_raw}"
            elif ordinal:
                suffix, label = f"S{ordinal:02d}", f"{kind_raw.strip()} {number_raw}"
            else:
                # 'X' prefix keeps prose-numbered places from ever colliding
                # with numerically-numbered ones (Cabinet 1 vs the first
                # prose-described cabinet).
                unnamed_seq[(room, kind)] += 1
                seq = unnamed_seq[(room, kind)]
                suffix = f"X{seq:02d}"
                label = (number_raw if number_raw and number_raw.upper() != "NA"
                         else f"{kind_raw.strip()} {seq}")
            location_id = f"{room}-{code}-{suffix}"

            # Every row in this sheet is a distinct physical place, so an id
            # collision here means two rows collapsed -- disambiguate loudly
            # rather than silently merging their contents.
            if location_id in self.locations:
                seq = 2
                while f"{location_id}-{seq}" in self.locations:
                    seq += 1
                self.warnings.append(
                    (sheet, rownum,
                     f"duplicate location id {location_id} -> {location_id}-{seq}"))
                location_id = f"{location_id}-{seq}"

            self.location(location_id, room, kind, number_raw, "",
                          label.strip() or location_id)

            if not contents_raw:
                loc = self.locations[location_id]
                if not loc["notes"]:
                    loc["notes"] = "no contents recorded in legacy spreadsheet"
                continue

            self.explode(contents_raw, location_id, sheet, rownum)

    def shelved(self, sheet: str, room: str, unit_kind: str, unit_id: str,
                unit_label: str, date_col: int | None = None,
                notes_col: int | None = None) -> None:
        """Sheets shaped: Shelf # / Contents / [Date] / [Notes]."""
        self.location(unit_id, room, unit_kind, "", "", unit_label)
        for rownum, values in self.rows(sheet):
            shelf_raw = clean(values[0])
            contents = clean(values[1]) if len(values) > 1 else ""
            if not contents:
                continue
            # Numeric/ordinal shelves become S01..S10 (zero-padded so they sort
            # correctly as strings). Named shelves ('Door Shelf 2') keep their
            # slug without the S prefix, which would read as 'SDOOR-SHELF-2'.
            ordinal = (int(shelf_raw) if shelf_raw.isdigit()
                       else ORDINALS.get(shelf_raw.lower()))
            if ordinal:
                location_id = f"{unit_id}-S{ordinal:02d}"
            elif shelf_raw:
                location_id = f"{unit_id}-{slug(shelf_raw)}"
            else:
                location_id = unit_id
            self.location(location_id, room, "shelf", shelf_raw, unit_id,
                          f"{unit_label} shelf {shelf_raw}")

            received, residue = "", ""
            if date_col is not None and len(values) > date_col:
                received, residue = parse_date(values[date_col], sheet, rownum,
                                               self.warnings)
            notes = ""
            if notes_col is not None and len(values) > notes_col:
                notes = clean(values[notes_col])
            if residue:
                notes = f"{notes}; {residue}".strip("; ") if notes else residue
            self.explode(contents, location_id, sheet, rownum, notes, received)

    def freezer_drawers(self, sheet: str, room: str, unit_id: str,
                        unit_label: str) -> None:
        """Shelf Number / Drawer Number / Contents."""
        self.location(unit_id, room, "freezer", "", "", unit_label)
        for rownum, values in self.rows(sheet):
            shelf_raw, drawer_raw = clean(values[0]), clean(values[1])
            contents = clean(values[2]) if len(values) > 2 else ""
            if not contents:
                continue
            shelf_no = f"{int(shelf_raw):02d}" if shelf_raw.isdigit() else slug(shelf_raw)
            shelf_id = self.location(
                f"{unit_id}-S{shelf_no}", room, "shelf", shelf_raw, unit_id,
                f"{unit_label} shelf {shelf_raw}")
            if drawer_raw and drawer_raw.upper() != "NA":
                drawer_no = (f"{int(drawer_raw):02d}" if drawer_raw.isdigit()
                             else slug(drawer_raw))
                location_id = self.location(
                    f"{shelf_id}-D{drawer_no}", room, "drawer", drawer_raw,
                    shelf_id, f"{unit_label} shelf {shelf_raw} drawer {drawer_raw}")
            else:
                location_id = shelf_id
            self.explode(contents, location_id, sheet, rownum)

    # -- driver -------------------------------------------------------------

    def run(self) -> None:
        names = self.wb.sheetnames
        pick = lambda frag: next(n for n in names if frag in n)

        self.drawers_cabinets(pick("DrawersCabinets"))
        self.shelved(pick("Refrigerator Rm 209"), "209", "refrigerator",
                     "209-FRIDGE", "Refrigerator Rm 209", date_col=2)
        self.shelved(pick("Shelf Rm 209"), "209", "shelf", "209-SHELF",
                     "Shelf unit Rm 209", date_col=2)
        self.shelved(pick("Shelf RM 213"), "213", "shelf", "213-SHELF",
                     "Shelves left of entryway Rm 213", date_col=2, notes_col=3)
        self.freezer_drawers(pick("-20"), "213", "213-F20", "-20C freezer Rm 213")

        self.write()
        self.report()

    def write(self) -> None:
        self.outdir.mkdir(parents=True, exist_ok=True)

        def dump(filename: str, fields: list[str], rows: list[dict]) -> None:
            with (self.outdir / filename).open("w", newline="") as handle:
                writer = csv.DictWriter(handle, fieldnames=fields)
                writer.writeheader()
                writer.writerows(rows)

        dump("locations.csv", LOCATION_FIELDS,
             sorted(self.locations.values(), key=lambda r: r["location_id"]))
        dump("items.csv", ITEM_FIELDS, self.items)
        dump("review_queue.csv", REVIEW_FIELDS, self.review)

    def report(self) -> None:
        print(f"locations     {len(self.locations):>5}")
        print(f"items         {len(self.items):>5}")
        print(f"review_queue  {len(self.review):>5}")
        print(f"skipped rows  {self.skipped:>5}  (cross-references only)")
        print("\ncategories:")
        for category, count in Counter(
                i["category"] for i in self.items).most_common():
            print(f"  {category:<12} {count:>4}")
        print("\nreview reasons:")
        for reason, count in Counter(
                r["reason"] for r in self.review).most_common():
            print(f"  {reason:<34} {count:>4}")
        if self.warnings:
            print(f"\ndata warnings ({len(self.warnings)}):")
            for sheet, rownum, message in self.warnings:
                print(f"  {sheet}:{rownum}  {message}")


def main() -> None:
    root = Path(__file__).resolve().parent.parent
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx",
                        default=root / "legacy" / "Roberts Lab Inventory.xlsx",
                        type=Path)
    parser.add_argument("--out", default=root / "data", type=Path)
    args = parser.parse_args()
    Migration(args.xlsx, args.out).run()


if __name__ == "__main__":
    main()
